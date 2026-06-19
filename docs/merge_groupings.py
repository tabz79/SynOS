import os
import openpyxl

def main():
    docs_dir = r"d:\Projects\SynOS-Synthesized-Lab-Intelligence\docs"
    enriched_file = os.path.join(docs_dir, "investigations_fixed_enriched.xlsx")
    group_file = os.path.join(docs_dir, "Exception_Grouping.xlsx")
    final_file = os.path.join(docs_dir, "investigations_fixed_final.xlsx")

    # Load Exception Grouping mappings into a dictionary
    wb_grp = openpyxl.load_workbook(group_file, data_only=True)
    ws_grp = wb_grp.active
    grp_rows = list(ws_grp.iter_rows(values_only=True))[1:]

    mappings = {}
    for r in grp_rows:
        if not r or r[1] is None:
            continue
        dept = str(r[1]).strip().upper()
        test_name = str(r[2]).strip().upper()
        param_name = str(r[4]).strip().upper() if r[4] is not None else ""
        spec = r[8]
        tube = r[9]
        key = (dept, test_name, param_name)
        mappings[key] = (spec, tube)

    print(f"Loaded {len(mappings)} mappings from Exception_Grouping.xlsx")

    # Load Enriched Workbook
    wb_enr = openpyxl.load_workbook(enriched_file)
    ws_enr1 = wb_enr["Sheet1"]
    
    # Update Sheet1 rows
    updated_count = 0
    # Header row is 1, data starts at 2
    for row_idx in range(2, ws_enr1.max_row + 1):
        dept_val = ws_enr1.cell(row=row_idx, column=1).value
        test_val = ws_enr1.cell(row=row_idx, column=2).value
        param_val = ws_enr1.cell(row=row_idx, column=4).value
        spec_cell = ws_enr1.cell(row=row_idx, column=8)
        tube_cell = ws_enr1.cell(row=row_idx, column=9)

        if dept_val is None:
            continue

        dept = str(dept_val).strip().upper()
        test_name = str(test_val).strip().upper() if test_val is not None else ""
        param_name = str(param_val).strip().upper() if param_val is not None else ""

        # Only update if Specimen or Tube is empty/None
        if spec_cell.value is None or tube_cell.value is None:
            key = (dept, test_name, param_name)
            if key in mappings:
                spec, tube = mappings[key]
                spec_cell.value = spec
                tube_cell.value = tube
                updated_count += 1
            else:
                # Try fallback matching on just department and test_name
                # (in case parameter name normalization had minor mismatches)
                fallback_keys = [k for k in mappings.keys() if k[0] == dept and k[1] == test_name]
                if fallback_keys:
                    spec, tube = mappings[fallback_keys[0]]
                    spec_cell.value = spec
                    tube_cell.value = tube
                    updated_count += 1

    print(f"Updated {updated_count} cells in Sheet1 with suggested values")

    # Save to final workbook
    wb_enr.save(final_file)
    print(f"Saved final workbook to {final_file}")

    # Double check if any are still blank
    blank_count = 0
    for row_idx in range(2, ws_enr1.max_row + 1):
        dept_val = ws_enr1.cell(row=row_idx, column=1).value
        spec_val = ws_enr1.cell(row=row_idx, column=8).value
        tube_val = ws_enr1.cell(row=row_idx, column=9).value
        if dept_val is not None and (spec_val is None or tube_val is None):
            blank_count += 1

    print(f"Verify: Remaining rows with blank Specimen/Tube: {blank_count}")

if __name__ == "__main__":
    main()
